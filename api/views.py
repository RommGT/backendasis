import os
import numpy as np
import cv2
from deepface import DeepFace
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.conf import settings
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

def decode_image(image_file):
    # Read the image from the uploaded file
    image_data = image_file.read()
    image_np = np.frombuffer(image_data, dtype=np.uint8)
    image = cv2.imdecode(image_np, cv2.IMREAD_COLOR)
    return image

def register_to_excel(email, class_name):
    EXCEL_FILE = settings.EXCEL_FILE
    # Load the Excel file if it exists
    if os.path.exists(EXCEL_FILE):
        df = pd.read_excel(EXCEL_FILE)
    else:
        df = pd.DataFrame(columns=['Correo', 'Curso', 'Fecha'])

    # Format the date and time
    timestamp = datetime.now().strftime('%d-%m-%Y %H:%M:%S')

    # Create a new entry
    new_entry = pd.DataFrame([{'Correo': email, 'Curso': class_name, 'Fecha': timestamp}])

    # Concatenate the new entry with the existing DataFrame
    df = pd.concat([df, new_entry], ignore_index=True)

    # Save the file temporarily in Excel to apply formats
    df.to_excel(EXCEL_FILE, index=False, header=True)  # Include headers

    # Load the file with openpyxl to apply formatting
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Format column headers
    headers = ['Correo', 'Curso', 'Fecha']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = PatternFill(start_color="16A24F", end_color="16A24F", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25

    # Save changes
    wb.save(EXCEL_FILE)

@csrf_exempt
@require_POST
def authenticate(request):
    if 'image' not in request.FILES:
        return JsonResponse({"error": "No image file provided"}, status=400)

    if 'email' not in request.POST:
        return JsonResponse({"error": "No email provided"}, status=400)

    if 'class' not in request.POST:
        return JsonResponse({"error": "No class provided"}, status=400)

    image_file = request.FILES['image']
    user_email = request.POST['email']
    class_name = request.POST['class']
    
    if not image_file or not user_email or not class_name:
        return JsonResponse({"error": "No image file, email, or class provided"}, status=400)

    # Decode the image from the file
    img = decode_image(image_file)

    # Temporarily save the uploaded image for comparison
    temp_image_path = os.path.join(settings.MEDIA_ROOT, f'{user_email}_temp.png')
    cv2.imwrite(temp_image_path, img)

    # Initialize the result as unknown
    best_match_email = "No se reconoce al usuario"
    best_match_score = float('inf')  # Lower score indicates better match

    # User's image directory
    user_folder_path = os.path.join(settings.IMAGENES_DIR, user_email)

    # Check if the user's folder exists
    if os.path.isdir(user_folder_path):
        for image_file_name in os.listdir(user_folder_path):
            image_path = os.path.join(user_folder_path, image_file_name)

            try:
                # Anti-spoofing test in face detection
                face_objs = DeepFace.extract_faces(
                    img_path=temp_image_path,
                    anti_spoofing=True
                )
                # Verify if all faces are real
                all_faces_real = all(face_obj["is_real"] for face_obj in face_objs)

                if not all_faces_real:
                    return JsonResponse({"error": "La cara no es real"})

                # Compare the uploaded image with the image in the user's folder
                result = DeepFace.verify(temp_image_path, image_path)
                
                # Assuming the result has a 'distance' key for the score
                score = result['distance']
                
                # Update the best result if it's better
                if score < best_match_score:
                    best_match_score = score
                    best_match_email = user_email

            except Exception as e:
                print(f"No se identificó el usuario: {e}")
    else:
        return JsonResponse({"error": "El usuario no existe"})

    # Clean up temporary file
    if os.path.exists(temp_image_path):
        os.remove(temp_image_path)

    if best_match_score < 0.55:
        # Register the information in the Excel file
        register_to_excel(best_match_email, class_name)
        return JsonResponse({"email": best_match_email})
    else:
        return JsonResponse({"error": "No se reconoce al usuario"})

@csrf_exempt
@require_POST
def register(request):
    if 'email' not in request.POST:
        return JsonResponse({"error": "Falta email"}, status=400)

    if 'images' not in request.FILES:
        return JsonResponse({"error": "Falta imagen"}, status=400)

    user_email = request.POST['email']
    images = request.FILES.getlist('images')

    if not user_email or not images:
        return JsonResponse({"error": "No hay imagen ni email"}, status=400)

    # Create a folder for the email if it doesn't exist
    user_folder_path = os.path.join(settings.IMAGENES_DIR, user_email)
    if not os.path.exists(user_folder_path):
        os.makedirs(user_folder_path)

    # Save each image in JPG format in the corresponding folder
    for index, image_file in enumerate(images):
        # Read the image from the uploaded file
        image_data = image_file.read()
        image_np = np.frombuffer(image_data, dtype=np.uint8)
        image = cv2.imdecode(image_np, cv2.IMREAD_COLOR)

        # Generate the filename based on an index
        image_filename = f'foto{index + 1}.jpg'
        image_path = os.path.join(user_folder_path, image_filename)

        # Save the image in JPG format
        cv2.imwrite(image_path, image)

    return JsonResponse({"result": "ok"}, status=200)

@require_GET
def get_attendance(request):
    EXCEL_FILE = settings.EXCEL_FILE

    if not os.path.exists(EXCEL_FILE):
        return JsonResponse({"error": "No existen registros"}, status=404)

    try:
        df = pd.read_excel(EXCEL_FILE)
        
        # Remove rows and columns with NaN values
        df = df.dropna(how='all')  # Remove completely empty rows
        df = df.dropna(axis=1, how='all')  # Remove completely empty columns
        
        # Convert to JSON
        attendance_json = df.to_dict(orient='records')
        return JsonResponse(attendance_json, safe=False, status=200)
    except Exception as e:
        return JsonResponse({"error": f"Error al leer el archivo Excel: {e}"}, status=500)
